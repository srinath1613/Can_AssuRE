import os
import shutil
import tkinter as tk
import tkinter.messagebox as tkmsg
from tkinter import ttk
import cantools
from tkinter import filedialog
import re
import importlib
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import config
import backend
from backend import ui_queue
import threading

# -------------------
# Global GUI root
# -------------------
root = tk.Tk()
header_frame_global = None
search_icon_global = None


# globals used across functions
tree = None
tree_frame = None
second_frame_global = None
connect_button_status_global = None
send = None

# Search UI globals
search_bar = None
search_entry = None
current_results = []
current_index = -1

# -----------------------
# Utility: update config file at runtime
# -----------------------
def update_config_file_runtime(variable_name, variable_value):
    """
    Safely updates a variable in config.py on disk.
    Keeps a readable representation and avoids extra brackets.
    """
    config_path = "config.py"

    with open(config_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    variable_found = False
    start_index = end_index = None

    for i, line in enumerate(lines):
        if re.match(rf'^\s*{variable_name}\s*=', line):
            variable_found = True
            start_index = i
            # detect block end
            end_index = i + 1
            open_brackets = 0
            while end_index < len(lines):
                line_stripped = lines[end_index].strip()
                open_brackets += line_stripped.count("[") + line_stripped.count("{")
                open_brackets -= line_stripped.count("]") + line_stripped.count("}")
                if open_brackets <= 0 and (re.match(r'^\s*\w+\s*=', line_stripped) or line_stripped == ""):
                    break
                end_index += 1
            if end_index >= len(lines):
                end_index = len(lines)
            break

    new_lines = []
    indent = "    "

    if isinstance(variable_value, list) and all(isinstance(item, dict) for item in variable_value):
        new_lines.append(f"{variable_name} = [\n")
        for idx, item in enumerate(variable_value):
            new_lines.append(indent + "{\n")
            for k, v in item.items():
                if k == "signals" and isinstance(v, dict):
                    new_lines.append(f"{indent * 2}'signals': {{\n")
                    for sk, sv in v.items():
                        new_lines.append(f"{indent * 3}'{sk}': {sv},\n")
                    new_lines.append(f"{indent * 2}}},\n")
                else:
                    new_lines.append(f"{indent * 2}'{k}': {repr(v)},\n")
            new_lines.append(indent + "}" + ("," if idx < len(variable_value) - 1 else "") + "\n")
        new_lines.append("]\n")

    elif isinstance(variable_value, dict):
        new_lines.append(f"{variable_name} = {{\n")
        for k, v in variable_value.items():
            new_lines.append(f"{indent}'{k}': {repr(v)},\n")
        new_lines.append("}\n")

    elif isinstance(variable_value, (int, float, bool)):
        new_lines.append(f"{variable_name} = {variable_value}\n")
    elif isinstance(variable_value, str):
        safe_path = variable_value.replace("\\", "/")
        new_lines.append(f'{variable_name} = r"{safe_path}"\n')
        #new_lines.append(f'{variable_name} = "{variable_value}"\n')
    else:
        new_lines.append(f"{variable_name} = {repr(variable_value)}\n")

    if variable_found:
        if start_index > 0 and lines[start_index - 1].strip().startswith("#"):
            start_index -= 1
        lines[start_index:end_index] = new_lines
    else:
        if not lines or not lines[-1].endswith("\n"):
            lines.append("\n")
        lines.extend(["\n"] + new_lines)

    with open(config_path, "w", encoding="utf-8") as f:
        f.writelines(lines)


# -----------------------
# Browse / load DBC + vector sheet
# -----------------------
def browse_files(messagebox, show_btn, send_btn, auto_send_btn, validate_btn, second_frame):
    """
    Let user pick a DBC file. Copies matching vector_list.xlsx if present,
    updates config, and refreshes backend caches.
    """
    filename = filedialog.askopenfilename(
        initialdir="./Files",
        title="Select File",
        filetypes=(("dbc files", "*.dbc"), ("All files", "*.*")),
    )

    if not filename:
        return
    # Reject selection inside the Mandatory folder
    mandatory_path = os.path.abspath(os.path.join("Files", "Mandatory"))
    selected_path = os.path.abspath(filename)

    if selected_path.startswith(mandatory_path):
        tkmsg.showerror(
            "Invalid Selection",
            "❌ Permission Denied to select full dbc file.\n"
            "Please choose listed dbc or create new dbc file."
        )
        return

    # show file in messagebox (read-only style)
    messagebox.config(state="normal")
    messagebox.delete("1.0", tk.END)
    messagebox.insert(tk.END, filename)
    messagebox.config(state="disabled")

    # update config and backend
    update_config_file_runtime("dbc_file_path", filename)
    importlib.reload(config)

    try:
        dbc_name = os.path.splitext(os.path.basename(filename))[0]
        excel_source = os.path.join("Files", f"{dbc_name}_vector_list.xlsx")
        excel_dest = os.path.join(os.getcwd(), "vector_list.xlsx")
        if os.path.exists(excel_source):
            shutil.copy2(excel_source, excel_dest)
            print(f"✅ Copied matching vector sheet: {excel_source} → {excel_dest}")
        else:
            print(f"⚠️ No matching Excel found for {dbc_name}. Expected: {excel_source}")
    except Exception as e:
        print(f"❌ Error copying vector sheet: {e}")

    # Update Vehicle property type mapping if applicable
    try:
        vehicle_type_dict = backend.update_vehicle_property_type(filename)
        update_config_file_runtime("Vehicle_propID_type", vehicle_type_dict)
        backend.copy_original_heartbeat_signal()
    except Exception as e:
        print(f"⚠️ vehicle property update failed: {e}")

    # Enable buttons
    show_btn.config(state=tk.NORMAL)
    send_btn.config(state=tk.NORMAL)
    auto_send_btn.config(state=tk.NORMAL)
    validate_btn.config(state=tk.NORMAL)

    # Refresh backend caches and preload vector sheet (clear caches first)
    try:
        # stop adb worker to avoid race
        try:
            backend.stop_adb_worker(timeout=2.0)
        except Exception:
            pass

        # clear backend caches so refresh picks up new files
        try:
            backend.cached_signal_dict.cache_clear()
        except Exception:
            pass
        try:
            backend.cached_property_map.cache_clear()
        except Exception:
            pass
        try:
            backend.cached_canid_to_signalname.cache_clear()
        except Exception:
            pass
        try:
            backend.cached_property_list.cache_clear()
        except Exception:
            pass
        try:
            backend.REGEX_CACHE.clear()
        except Exception:
            pass
        try:
            backend.EXCEL_FAST_CACHE.clear()
            backend.EXCEL_FAST_CACHE.update({
                "wb": None,
                "sheet": None,
                "time_row": {},
                "tx_col": {},
                "rx_col": {},
                "initialized": False,
                "tx_initialized": False
            })
        except Exception:
            pass
        try:
            backend.frameid_map.clear()
            backend._frameid_map_loaded = False
        except Exception:
            pass

        # Now refresh config and preload vector sheet
        backend.refresh_config_and_dbc()
        backend.preload_vector_sheet()
    except Exception as e:
        print(f"⚠️ backend refresh failed: {e}")

    update_ui_from_queue()


# -----------------------
# Clear UI and state
# -----------------------
def clear_text(messagebox, show_btn, send_btn, auto_send_btn, browse_button, validate_btn):
    messagebox.config(state="normal")
    messagebox.delete("1.0", tk.END)
    messagebox.config(state="disabled")
    show_btn.config(state="disabled")
    auto_send_btn.config(state="disabled")
    send_btn.config(state="disabled")
    validate_btn.config(state="disabled")
    browse_button.config(state="normal")

    # destroy tree/table if present
    global tree_frame, tree
    try:
        if tree_frame:
            tree_frame.destroy()
            tree_frame = None
            tree = None
    except Exception:
        pass

    # clear backend queues & config signals
    backend.user_signal_queue.queue.clear()
    backend.ui_queue.queue.clear()
    update_config_file_runtime("user_send_signals", [])
    importlib.reload(config)
    backend.time_column_initialized = False
    try:
        update_config_file_runtime("user_send_signals", [])
        update_config_file_runtime("Vehicle_propID_type", {})
        update_config_file_runtime("heart_beat_signals", list(backend.original_heartbeat_backup))
        importlib.reload(config)
    except Exception as e:
        print(f"⚠️ Failed to clear user_send_signals: {e}")
    # ======================================================
    # STOP ADB BACKGROUND WORKER (wait) + RESET BACKEND CACHES
    # ======================================================
    try:
        # ask backend to stop the adb worker and wait (backend.stop_adb_worker joins)
        backend.stop_adb_worker(timeout=2.0)
        print("🛑 Requested backend ADB worker stop")
    except Exception as e:
        print("⚠️ Failed requesting stop to ADB worker:", e)

    # reset last cached ADB values
    try:
        with getattr(backend, "shared_rx_lock"):
            backend.shared_rx_latest = {}
    except Exception:
        try:
            backend.shared_rx_latest = {}
        except:
            pass

    # stop synchronized worker too
    try:
        backend.stop_synchronized_worker()
    except Exception as e:
        print("⚠️ Failed stopping synchronized worker:", e)

    # Clear all backend in-memory caches the worker uses
    try:
        # clear lru_cache wrapped functions (safe to call even if not cached)
        try:
            backend.cached_signal_dict.cache_clear()
        except Exception:
            pass
        try:
            backend.cached_property_map.cache_clear()
        except Exception:
            pass
        try:
            backend.cached_canid_to_signalname.cache_clear()
        except Exception:
            pass
        try:
            backend.cached_property_list.cache_clear()
        except Exception:
            pass

        # clear fast-excel memory
        try:
            backend.EXCEL_FAST_CACHE.clear()
            backend.EXCEL_FAST_CACHE.update({
                "wb": None,
                "sheet": None,
                "time_row": {},
                "tx_col": {},
                "rx_col": {},
                "initialized": False,
                "tx_initialized": False
            })
        except Exception:
            pass

        # clear other maps used by worker
        try:
            backend.frameid_map.clear()
        except Exception:
            pass
        try:
            backend._frameid_map_loaded = False
        except Exception:
            pass

        # clear regex cache and other caches
        try:
            backend.REGEX_CACHE.clear()
        except Exception:
            pass

        print("🔁 All backend caches cleared")
    except Exception as e:
        print("⚠️ Failed clearing backend cache:", e)
    send_btn.config(state="disabled")


def select_parent_dbc_and_continue(messagebox=None, show=None, send=None, auto_send=None, browse_button=None):
    """
    Step 1 → User selects parent DBC file (checkbox list)
    Step 2 → Saves to config.full_frame_dbc
    Step 3 → Calls existing generate_dbc() function
    """
    mandatory_path = os.path.join(os.getcwd(), "Files", "Mandatory")
    dbc_files = [f for f in os.listdir(mandatory_path) if f.lower().endswith(".dbc")]

    if not dbc_files:
        tkmsg.showerror("Error", "❌ No DBC files found inside Files/Mandatory/")
        return

    # -----------------------------
    # Popup to choose parent DBC
    # -----------------------------
    popup = tk.Toplevel()
    popup.title("Select Parent DBC")
    popup.geometry("450x500")
    popup.grab_set()

    tk.Label(popup, text="Select Parent DBC File:",
             font=("Arial", 11, "bold")).pack(pady=10)

    # ---- scrollable frame ----
    container = tk.Frame(popup)
    canvas = tk.Canvas(container, height=350)
    scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>",
                          lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    container.pack(fill="both", expand=True)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Only one checkbox allowed
    parent_vars = {}
    selected_parent = tk.StringVar(value="")

    def checkbox_clicked(selected):
        selected_parent.set(selected)
        # uncheck others
        for name, var in parent_vars.items():
            if name != selected:
                var.set(False)

    # -- Add checkboxes for each parent file --
    for f in dbc_files:
        var = tk.BooleanVar()
        cb = tk.Checkbutton(scrollable_frame, text=f, variable=var,
                            command=lambda name=f: checkbox_clicked(name),
                            anchor="w", justify="left")
        cb.pack(fill="x", padx=15)
        parent_vars[f] = var

    # Choose button
    choose_btn = tk.Button(
        popup, text="Continue",
        state="disabled",
        bg="#4CAF50", fg="white",
        width=25
    )
    choose_btn.pack(pady=20)

    # Enable button only when selection is made
    def update_button(*_):
        choose_btn.config(state="normal" if selected_parent.get() else "disabled")

    selected_parent.trace_add("write", update_button)

    # ---- Finish selection ----
    def finish():
        chosen = selected_parent.get()
        selected_path = os.path.join(mandatory_path, chosen)

        # Save in config file
        update_config_file_runtime("full_frame_dbc", selected_path)
        importlib.reload(config)

        popup.destroy()

        # Now call your ORIGINAL function
        generate_dbc(messagebox=None, show=None, send=None, auto_send=None, browse_button=None)

    choose_btn.config(command=finish)


def generate_dbc(messagebox=None, show=None, send=None, auto_send=None, browse_button=None):
    """Popup to select frames from full DBC and create a new smaller DBC."""

    # --- Load full DBC file ---
    try:
        if config.Testing_device == "Linux":
            dbc_path = os.path.join(os.getcwd(), config.full_frame_dbc)
        else:
            dbc_path = config.full_frame_dbc
        db = cantools.database.load_file(dbc_path)
    except Exception as e:
        messagebox.showerror("Error", f"❌ Failed to load full DBC file:\n{e}")
        return

    # --- Create popup ---
    popup = tk.Toplevel()
    popup.title("Select Frames to Create New DBC")
    popup.geometry("550x600")
    popup.grab_set()  # make modal

    # --- Scrollable area ---
    container = tk.Frame(popup)
    canvas = tk.Canvas(container)
    scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    container.pack(fill="both", expand=True)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # --- Mouse scroll support ---
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    canvas.bind_all("<MouseWheel>", _on_mousewheel)  # for Windows
    canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux up
    canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))   # Linux down

    # --- Title ---
    tk.Label(
        scrollable_frame,
        text="Select Frames to include in new DBC:",
        font=("Arial", 11, "bold")
    ).pack(pady=8)

    # --- Checkbox for each frame ---
    frame_vars = {}
    for msg in db.messages:
        var = tk.BooleanVar()
        label_text = f"{msg.name} (0x{msg.frame_id:X})"
        cb = tk.Checkbutton(scrollable_frame, text=label_text, variable=var, anchor="w", justify="left")
        cb.pack(fill="x", padx=15)
        frame_vars[msg.name] = var
    # --- Select All / Clear All for frames ---
    btn_frame2 = tk.Frame(scrollable_frame)
    btn_frame2.pack(pady=10)

    def select_all_frames():
        for var in frame_vars.values():
            var.set(True)

    def clear_all_frames():
        for var in frame_vars.values():
            var.set(False)

    tk.Button(btn_frame2, text="Select All", width=12, command=select_all_frames).grid(row=0, column=0, padx=5)
    tk.Button(btn_frame2, text="Clear All", width=12, command=clear_all_frames).grid(row=0, column=1, padx=5)

    # --- Output filename ---
    tk.Label(
        scrollable_frame,
        text="\nEnter output DBC filename:",
        font=("Arial", 10, "bold")
    ).pack(pady=(10, 2))

    filename_entry = tk.Entry(scrollable_frame, width=40)
    filename_entry.pack(pady=5)
    filename_entry.insert(0, "selected_frames.dbc")

    # --- Save button (initially disabled) ---
    save_btn = tk.Button(
        scrollable_frame,
        text="Save Selected Frames",
        state="disabled",
        bg="#4CAF50", fg="white",
        width=25, height=1
    )
    save_btn.pack(pady=15)

    # --- Check if any frame selected ---
    def update_button_state(*_):
        if any(var.get() for var in frame_vars.values()):
            save_btn.config(state="normal")
        else:
            save_btn.config(state="disabled")

    for var in frame_vars.values():
        var.trace_add("write", update_button_state)

    # --- Save logic ---

    def save_selected():
        selected_frames = [name for name, var in frame_vars.items() if var.get()]
        save_name = filename_entry.get().strip()
        if not save_name.endswith(".dbc"):
            save_name += ".dbc"

        files_dir = os.path.join(os.getcwd(), "Files")
        os.makedirs(files_dir, exist_ok=True)
        save_path = os.path.join(files_dir, save_name)

        try:
            # ✅ Load full DBC text
            with open(dbc_path, "r", encoding="utf-8", errors="ignore") as full_file:
                lines = full_file.readlines()

            selected_msgs = [m for m in db.messages if m.name in selected_frames]
            selected_names = [m.name for m in selected_msgs]

            new_dbc_lines = []
            copying = False
            frame_name = None

            for line in lines:
                if line.startswith("BO_ "):
                    frame_name = line.split(":")[0].split()[-1]
                    copying = frame_name in selected_names

                if copying:
                    new_dbc_lines.append(line)

                # Detect next frame start and end previous
                if copying and line.strip() == "":
                    new_dbc_lines.append("\n")
                    copying = False

            # ✅ Write header + selected frames to new file
            with open(save_path, "w", encoding="utf-8") as out_file:
                # Copy header portion until first BO_
                for line in lines:
                    if line.startswith("BO_ "):
                        break
                    out_file.write(line)
                # Write selected frame definitions
                for line in new_dbc_lines:
                    out_file.write(line)

            tkmsg.showinfo("Success", f"✅ DBC file saved successfully to:\n{save_path}")

            # ✅ Generate Excel using the newly created DBC
            DBC_PATH = save_path
            base_name = os.path.splitext(os.path.basename(save_name))[0]
            OUTPUT_XLSX = os.path.join(files_dir, f"{base_name}_vector_list.xlsx")

            TYPE_H_PATH = os.path.join(os.getcwd(), config.type_h_file)

            if os.path.exists(TYPE_H_PATH):
                create_excel_sheet(TYPE_H_PATH, DBC_PATH, OUTPUT_XLSX)
                tkmsg.showinfo("Excel Created", f"📘 Excel file generated:\n{OUTPUT_XLSX}")
            else:
                tkmsg.showwarning("Warning", f"⚠️ 'types.h' not found at:\n{TYPE_H_PATH}")

            popup.destroy()

        except Exception as e:
            tkmsg.showerror("Error", f"❌ Failed to save new DBC or create Excel:\n{e}")

    save_btn.config(command=save_selected)



# -----------------------
# Create Excel sheet from DBC (kept from your file)
# -----------------------
def create_excel_sheet(TYPE_H_PATH, DBC_PATH, OUTPUT_XLSX):
    # Build quick lookup for heartbeat values:
    heartbeat_lookup = {}  # { frame_name: {signal: value} }

    for hb in config.heart_beat_signals:
        fname = hb["frame_name"]
        heartbeat_lookup[fname] = hb["signals"]
    try:
        db = cantools.database.load_file(DBC_PATH)
        frame_names = {msg.name for msg in db.messages}
        frame_list = [{"Can_Messages": msg.name, "Can_ID": hex(msg.frame_id)} for msg in db.messages]

        pattern = re.compile(r"([A-Za-z0-9_]+)__([A-Za-z0-9_]+)_[RT]X_V\s*=\s*(\d+)", re.IGNORECASE)

        # --------------------------------------------------------------
        # STEP 1: Build FULL signal list from DBC
        # --------------------------------------------------------------
        dbc_signals = []  # temporary list containing all signals from DBC
        for msg in db.messages:
            for sig in msg.signals:
                dbc_signals.append({
                    "message": msg.name,
                    "signal": sig.name,
                    "property_id": "N/A"  # default if type.h not found
                })

        # --------------------------------------------------------------
        # STEP 2: Read TYPE_H file and build lookup
        # --------------------------------------------------------------
        typeh_lookup = {}
        with open(TYPE_H_PATH, "r", encoding="utf-8") as f:
            for line in f:
                match = pattern.search(line)
                if match:
                    msg_name = match.group(1)
                    signal_name = match.group(2)
                    prop_id_dec = int(match.group(3))
                    typeh_lookup[(msg_name, signal_name)] = hex(prop_id_dec)

        # --------------------------------------------------------------
        # STEP 3: Merge type.h values into full DBC signal list
        # --------------------------------------------------------------
        signals = []
        for entry in dbc_signals:
            key = (entry["message"], entry["signal"])
            entry["property_id"] = typeh_lookup.get(key, "N/A")
            signals.append(entry)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "CAN_Signals"
        ws1["A1"] = "Property ID"
        ws1["A2"] = "Time (ms)"

        for idx, sig in enumerate(signals, start=2):
            col = get_column_letter(idx)
            ws1[f"{col}1"] = sig["property_id"]
            ws1[f"{col}2"] = sig["signal"]

        time_values = [1000, 2000, 3000]
        for r, t in enumerate(time_values, start=3):
            ws1[f"A{r}"] = t
            for c, sig in enumerate(signals, start=2):
                col_letter = get_column_letter(c)

                frame_name = sig["message"]
                signal_name = sig["signal"]

                # Default Tx value
                tx_val = 0

                # If heartbeat frame → use heartbeat default value
                if frame_name in heartbeat_lookup:
                    if signal_name in heartbeat_lookup[frame_name]:
                        tx_val = heartbeat_lookup[frame_name][signal_name]

                ws1[f"{col_letter}{r}"] = tx_val

        ws2 = wb.create_sheet("FrameID")
        ws2.append(["Can_Messages", "Can_ID"])
        for frame in frame_list:
            ws2.append([frame["Can_Messages"], frame["Can_ID"]])

        ws3 = wb.create_sheet("Results")
        ws3["A1"] = "Property ID"
        ws3["A2"] = "Time (ms)"

        col_idx = 2
        for sig in signals:
            start_col = col_idx
            end_col = col_idx + 1
            ws3.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws3.cell(row=1, column=start_col, value=sig["property_id"])
            ws3.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            ws3.cell(row=2, column=start_col, value=sig["signal"])
            ws3.cell(row=3, column=start_col, value="Tx")
            ws3.cell(row=3, column=end_col, value="Rx Car Service")
            col_idx += 2

        wb.save(OUTPUT_XLSX)
        print(f"✅ Excel file generated successfully: {OUTPUT_XLSX}")
        return True
    except Exception as e:
        print(f"❌ Failed to create Excel: {e}")
        return False


# -----------------------
# Treeview create_table (professional table)
# -----------------------
def create_table(second_frame, browse_button):
    """
    Build a professional Treeview table inside second_frame.
    Ensures backend caches are loaded before rendering.
    """
    global tree, tree_frame, second_frame_global

    # Force backend reload and preload vector sheet
    try:
        backend.check_auto_send_time_in_excel_update_config_file()
        backend._frameid_map_loaded = False
        backend.refresh_config_and_dbc()
        backend.preload_vector_sheet()
    except Exception as e:
        print(f"⚠️ Error preloading backend data: {e}")

    # Destroy previous table if exists
    try:
        if tree_frame:
            tree_frame.destroy()
    except Exception:
        pass

    tree_frame = tk.Frame(second_frame, bg="#FFFFFF")
    tree_frame.pack(fill=tk.BOTH, expand=1, padx=8, pady=8)

    # Style
    style = ttk.Style()
    try:
        style.theme_use("default")
    except Exception:
        pass
    style.configure("Treeview",
                    background="#FFFFFF",
                    fieldbackground="#FFFFFF",
                    foreground="#1A1A1A",
                    rowheight=28,
                    font=("Segoe UI", 10))

    style.configure("Treeview.Heading",
                    background="#3A6D8C",
                    foreground="white",
                    font=("Segoe UI", 10, "bold"))

    style.map("Treeview",
              background=[("selected", "#DDE8F0")],
              foreground=[("selected", "black")])

    style.configure("Vertical.TScrollbar", background="#2C3E50")
    style.configure("Horizontal.TScrollbar", background="#2C3E50")

    # Columns
    # NEW COLUMN ORDER
    columns = ("SNo", "Can_Messages", "Signals", "Property_Value", "Value_Tx", "Car_Service_Rx")
    tree_local = ttk.Treeview(tree_frame, columns=columns, show="headings")

    # Headings
    tree_local.heading("SNo", text="S.No", anchor="center")
    tree_local.heading("Can_Messages", text="CAN MESSAGES", anchor="w")
    tree_local.heading("Signals", text="SIGNALS", anchor="w")
    tree_local.heading("Property_Value", text="PROPERTY VALUE", anchor="w")
    tree_local.heading("Value_Tx", text="VALUE Tx", anchor="center")
    tree_local.heading("Car_Service_Rx", text="CAR SERVICE Rx", anchor="center")

    # Column widths and anchors
    tree_local.column("SNo", width=60, anchor="center", stretch=False)
    tree_local.column("Can_Messages", width=220, anchor="w", stretch=True)
    tree_local.column("Signals", width=300, anchor="w", stretch=True)
    tree_local.column("Property_Value", width=220, anchor="w", stretch=True)
    tree_local.column("Value_Tx", width=120, anchor="center", stretch=False)
    tree_local.column("Car_Service_Rx", width=160, anchor="center", stretch=False)

    # Scrollbars (VH)
    vscroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree_local.yview)
    hscroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree_local.xview)
    tree_local.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)

    vscroll.pack(side=tk.RIGHT, fill=tk.Y)
    hscroll.pack(side=tk.BOTTOM, fill=tk.X)
    tree_local.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

    # Populate rows from backend
    frame_list = []
    try:
        frame_list = backend.get_list_of_frames()
    except Exception as e:
        print(f"⚠️ get_list_of_frames failed: {e}")

    signals_data = []
    heartbeat_frame_map = {hb["frame_name"]: hb["signals"] for hb in config.heart_beat_signals}
    for frame_name in frame_list:
        try:
            sigs = backend.get_list_of_signals(frame_name)
        except Exception:
            sigs = []
        for s in sigs:
            # --- NEW LOGIC HERE ---
            if frame_name in heartbeat_frame_map:
                # real heartbeat value (may be 99, 10, 5 etc)
                val_tx = heartbeat_frame_map[frame_name].get(s, 0)
            else:
                # regular UI behaviour
                val_tx = "0" if getattr(config, "Testing_device", "") == "Linux" else ""
            signals_data.append({
                "frame": frame_name,
                "signal": s,
                "val_tx": val_tx,
                "property": "",
                "val_rx": ""
            })

    # Try to map property values if available
    try:
        prop_vals = backend.get_property_values()
    except Exception:
        prop_vals = []
    if len(prop_vals) == len(signals_data) and len(prop_vals) > 0:
        for i in range(len(signals_data)):
            signals_data[i]["property"] = prop_vals[i]
    else:
        for i in range(min(len(prop_vals), len(signals_data))):
            signals_data[i]["property"] = prop_vals[i]

    # Insert rows
    s_no = 1
    for row in signals_data:
        iid = f"{row['frame']}::{row['signal']}"
        tree_local.insert("", "end", iid=iid,
                          values=(s_no, row["frame"], row["signal"], row["property"], row["val_tx"], row["val_rx"]))
        s_no += 1
    # Inline editing for VALUE Tx only (column index 3)
    def on_double_click(event):
        region = tree_local.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tree_local.identify_column(event.x)  # like '#1'
        row_iid = tree_local.identify_row(event.y)
        if not row_iid:
            return
        col_index = int(col.replace("#", ""))
        # Only Value_Tx is editable -> that's column index 3 (1-based)
        if col_index != 5:
            return

        x, y, width, height = tree_local.bbox(row_iid, col)
        cur_val = tree_local.set(row_iid, column=columns[col_index - 1])

        entry = tk.Entry(tree_local, font=("Segoe UI", 10), justify="center")
        entry.insert(0, cur_val)
        entry.place(x=x, y=y, width=width, height=height)
        entry.focus_set()

        def save_edit(event=None):
            new_val = entry.get()
            tree_local.set(row_iid, column=columns[col_index - 1], value=new_val)
            entry.destroy()

        def cancel_edit(event=None):
            entry.destroy()

        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit)
        entry.bind("<Escape>", cancel_edit)

    tree_local.bind("<Double-1>", on_double_click)

    # store to globals
    globals()["tree"] = tree_local
    globals()["tree_frame"] = tree_frame
    globals()["second_frame_global"] = second_frame

    # disable browse button after table creation
    try:
        browse_button.config(state="disabled")
    except Exception:
        pass
    backend.start_adb_worker()
    return tree_local


# ---------------
# SEARCH BAR
# --------------

def close_search(tree):
    """Close the inline search bar and remove all highlights."""
    global search_bar, search_entry, current_results, current_index

    # Destroy bar if exists
    try:
        if search_bar and search_bar.winfo_exists():
            search_bar.destroy()
    except:
        pass

    search_bar = None
    search_entry = None
    current_results = []
    current_index = -1

    # Clear Treeview selection
    if tree:
        try:
            sel = tree.selection()
            if sel:
                tree.selection_remove(sel)
        except:
            pass


def open_search_inline(tree, header_frame, search_icon_btn):
    """Create an inline search bar inside header_frame."""
    global search_bar, search_entry, current_results, current_index

    if search_bar and search_bar.winfo_exists():
        close_search(tree)
        return

    current_results = []
    current_index = -1

    # create container
    search_bar = tk.Frame(header_frame, bg="#1F2B3C", highlightbackground="#FFFFFF", highlightthickness=1)
    header_frame.update_idletasks()

    # position left of search icon
    icon_x = search_icon_btn.winfo_x()
    icon_y = search_icon_btn.winfo_y()

    place_x = max(10, icon_x - 360)
    place_y = icon_y

    search_bar.place(x=place_x, y=place_y, width=360, height=36)

    # Entry box
    search_entry = tk.Entry(search_bar, font=("Segoe UI", 11))
    search_entry.pack(side="left", padx=6, pady=4, fill="x", expand=True)
    search_entry.focus_set()

    # Perform search when pressing Enter
    search_entry.bind("<Return>", lambda e: perform_search(tree))
    search_bar.bind_all("<Escape>", lambda e: close_search(tree))

    # Prev / Next buttons
    tk.Button(search_bar, text="◀", width=3,
              command=lambda: goto_prev(tree)).pack(side="left", padx=2, pady=3)

    tk.Button(search_bar, text="▶", width=3,
              command=lambda: goto_next(tree)).pack(side="left", padx=2, pady=3)

    # Close button
    tk.Button(search_bar, text="✖", width=3,
              command=lambda: close_search(tree)).pack(side="left", padx=4, pady=3)


def perform_search(tree):
    global search_entry, current_results, current_index

    if not search_entry:
        return

    query = search_entry.get().strip().lower()
    current_results = []
    current_index = -1

    if not query:
        return

    for iid in tree.get_children():
        row = tree.item(iid).get("values", [])
        if any(query in str(v).lower() for v in row):
            current_results.append(iid)

    if current_results:
        current_index = 0
        highlight(tree, current_results[0])


def highlight(tree, iid):
    try:
        tree.selection_set(iid)
        tree.see(iid)
    except:
        pass


def goto_next(tree):
    """Go to next search result."""
    global current_index, current_results
    if not current_results:
        return
    current_index = (current_index + 1) % len(current_results)
    highlight(tree, current_results[current_index])


def goto_prev(tree):
    """Go to previous search result."""
    global current_index, current_results
    if not current_results:
        return
    current_index = (current_index - 1) % len(current_results)
    highlight(tree, current_results[current_index])


# -----------------------
# Update functions used by backend -> update UI
# -----------------------
def update_tx_values(tx_dict):
    """
    tx_dict: { signal_name: value, ... } or nested {"tx_update": {...}}
    This updates the Treeview Value_Tx column by matching SIGNAL name.
    """
    if not tx_dict:
        return
    # tx_dict may come nested: {"tx_update": {sig: val, ...}}
    if "tx_update" in tx_dict and isinstance(tx_dict["tx_update"], dict):
        tx_entries = tx_dict["tx_update"]
    else:
        tx_entries = tx_dict

    try:
        if tree is None:
            return
        for sig_name, val in tx_entries.items():
            for iid in tree.get_children():
                if tree.set(iid, "Signals") == sig_name:
                    tree.set(iid, "Value_Tx", val)
                    break
    except Exception as e:
        print(f"⚠️ update_tx_values error: {e}")


def update_rx_values(rx_dict):
    """
    rx_dict: { signal_name: value, ... }
    Update CAR SERVICE Rx column (read-only for user).
    """
    if not rx_dict:
        return
    try:
        if tree is None:
            return
        for sig_name, val in rx_dict.items():
            for iid in tree.get_children():
                if tree.set(iid, "Signals") == sig_name:
                    tree.set(iid, "Car_Service_Rx", val)
                    break
    except Exception as e:
        print(f"⚠️ update_rx_values error: {e}")


# -----------------------
# UI queue polling & processing
# -----------------------
def update_ui_from_queue():
    try:
        while not ui_queue.empty():
            item = ui_queue.get_nowait()
            # handle peak device status
            if isinstance(item, dict) and "peak_status" in item:
                status = item["peak_status"]
                if status == "connected":
                    if connect_button_status_global:
                        connect_button_status_global.config(text="Connected", fg="#32CD32")
                else:
                    if connect_button_status_global:
                        connect_button_status_global.config(text="Error Connecting...", fg="red")
                continue

            # handle tx updates nested as {"tx_update": {...}}
            if isinstance(item, dict) and "tx_update" in item:
                try:
                    update_tx_values(item["tx_update"])
                except Exception as e:
                    print(f"⚠️ Failed to apply tx update: {e}")
                continue

            # otherwise assume rx mapping (signal_name -> value)
            if isinstance(item, dict):
                try:
                    update_rx_values(item)
                except Exception as e:
                    print(f"⚠️ Failed to apply rx update: {e}")

    except Exception as e:
        print(f"⚠️ UI update error: {e}")
    finally:
        root.after(100, update_ui_from_queue)


# -----------------------
# Auto send (delegates to backend)
# -----------------------
def send_signal_auto_from_list(second_frame):
    try:
        if not getattr(backend, "_synchronized_worker_started", False):
            backend.start_synchronized_worker()
            backend._synchronized_worker_started = True
            print("✅ Synchronized worker started.")
            # 🔥 Disable send button
            send.config(state=tk.DISABLED)
        else:
            print("⚙️ Worker already running.")
    except Exception as e:
        print(f"❌ Failed to start worker: {e}")


# -----------------------
# Read user values from Treeview and update config.user_send_signals
# -----------------------
def read_signal_value_from_ui(second_frame):
    """
    Read VALUE Tx values from the Treeview and build user_send_signals
    structure that backend heartbeat thread expects.
    """
    try:
        # ensure vector sheet loaded
        backend._frameid_map_loaded = False
        backend.preload_vector_sheet()

        if tree is None:
            print("⚠️ No table present to read from.")
            return

        # group by frame
        frame_groups = {}
        for iid in tree.get_children():
            vals = tree.item(iid, "values")
            frame_name = vals[1]
            sig_name = vals[2]
            try:
                tx_val = int(vals[4])
            except Exception:
                # try convert float -> int or default 0
                try:
                    tx_val = int(float(vals[4]))
                except Exception:
                    tx_val = 0
            frame_groups.setdefault(frame_name, {})[sig_name] = tx_val

        new_user_send_signals = []
        heartbeat_frames = {hb["frame_name"] for hb in config.heart_beat_signals}
        for frame_name, sig_map in frame_groups.items():
            frame_id = backend.load_vector_sheet_frameid(frame_name)
            if frame_name in heartbeat_frames and frame_id:
                for hb in config.heart_beat_signals:
                    if hb["frame_name"] == frame_name:
                        for sig_name, val in sig_map.items():
                            if sig_name in hb["signals"]:
                                hb["signals"][sig_name] = val
                        break
            if frame_name not in heartbeat_frames and frame_id:
                new_user_send_signals.append({
                    "frame_name": frame_name,
                    "can_id": frame_id,
                    "signals": sig_map
                })
        update_config_file_runtime("heart_beat_signals", config.heart_beat_signals)
        #update_config_file_runtime("user_send_signals", new_user_send_signals)
        importlib.reload(config)
        backend.user_send_signals_runtime = new_user_send_signals
        print("✅ user_send_signals updated from UI.")

        # Trigger backend validation (optional)
        adb_rx = backend.validate_vhal_layer()
        ui_queue.put(adb_rx)
    except Exception as e:
        print(f"❌ Error reading UI values: {e}")


# -----------------------
# Validate button
# -----------------------
def validate(second_frame):
    adb_rx = backend.validate_vhal_layer()
    ui_queue.put(adb_rx)


# -----------------------
# Main home_page (fixed header + table area)
# -----------------------
def home_page():
    root.title(getattr(config, "GUI_title", "CAN Tool"))
    root.geometry(getattr(config, "Display_size", "1000x700"))
    root.configure(bg="#F2F4F7")

    try:
        close_search(globals().get("tree", None))
    except:
        pass

    # Fixed header frame (top)
    header_frame = tk.Frame(root, bg="#2C3E50")
    header_frame.pack(fill=tk.X)


    # Main frame where Treeview will live
    main_frame = tk.Frame(root, bg="#F2F4F7")
    main_frame.pack(fill=tk.BOTH, expand=1)

    # second_frame is container for the table
    second_frame = tk.Frame(main_frame, bg="#F2F4F7")
    second_frame.pack(fill=tk.BOTH, expand=1, padx=6, pady=6)

    # --- Header widgets ---
    info_label0 = tk.Label(
        header_frame, text="Peak Device :",
        bg="#2C3E50", fg="#FFFFFF", font=("Segoe UI", 11, "bold")
    )
    info_label0.grid(row=0, column=0, pady=8, padx=8, sticky="w")

    # Peak device status label
    connect_status = tk.Label(header_frame, text="Checking...", bg="#2C3E50", fg="#FFFFFF", font=("Segoe UI", 11, "bold"))
    connect_status.grid(row=0, column=1, pady=8, padx=8, sticky="w")
    globals()["connect_button_status_global"] = connect_status

    info_label1 = tk.Label(header_frame, text="Load dbc files : ", bg="#2C3E50", fg="#FFFFFF", font=("Segoe UI", 11, "bold"))
    info_label1.grid(row=1, column=0, pady=8, padx=8, sticky="w")

    messagebox = tk.Text(header_frame, height=2, width=40, wrap="none")
    messagebox.grid(row=1, column=1, pady=8, padx=8, sticky="w")
    messagebox.config(state="disabled")


    show_btn = tk.Button(header_frame, text="Show Signals", state=tk.DISABLED,
                         command=lambda: create_table(second_frame, browse_btn))
    show_btn.grid(row=2, column=0, pady=8, padx=6)

    global send
    #send = tk.Button(
    #    header_frame, text="Send signal", state=tk.DISABLED,
    #    command=lambda: (close_search(tree), read_signal_value_from_ui(second_frame))
    #)
    send = tk.Button(
        header_frame, text="Send signal", state=tk.DISABLED,
        command=lambda: threading.Thread(
        target=lambda: read_signal_value_from_ui(second_frame),
        daemon=True
    ).start())

    send.grid(row=2, column=2, pady=8, padx=6)

    validate_btn = tk.Button(
        header_frame, text="Validate", state=tk.DISABLED,
        command=lambda: (close_search(tree), validate(second_frame))
    )
    validate_btn.grid(row=2, column=3, pady=8, padx=6)

    auto_send_btn = tk.Button(header_frame, text="Auto send", state=tk.DISABLED,
                              command=lambda: send_signal_auto_from_list(second_frame))
    auto_send_btn.grid(row=2, column=1, pady=8, padx=6)

    browse_btn = tk.Button(header_frame, text="Browse",
                           command=lambda: browse_files(messagebox, show_btn, send, auto_send_btn, validate_btn, second_frame))
    browse_btn.grid(row=1, column=2, pady=8, padx=6)

    clear_btn = tk.Button(header_frame, text="Clear all",
                          command=lambda: clear_text(messagebox, show_btn, send, auto_send_btn, browse_btn, validate_btn))
    clear_btn.grid(row=1, column=3, pady=8, padx=6)

    create_btn = tk.Button(header_frame, text="Create",
                           command=lambda: select_parent_dbc_and_continue(messagebox, show_btn, send, auto_send_btn, browse_btn))
    create_btn.grid(row=1, column=4, pady=8, padx=6)

    search_icon_btn = tk.Button(
        header_frame,
        text="🔍",
        font=("Segoe UI", 14),
        bg="#2C3E50",
        fg="white",
        relief="flat",
        command=lambda: open_search_inline(tree, header_frame, search_icon_btn)
    )
    search_icon_btn.grid(row=0, column=5, padx=8)  # adjust row/column if you want it top-right
    globals()["search_icon_global"] = search_icon_btn
    # Backend initial checks
    backend.check_peak_device_interface_name()
    backend.make_can_interface_up()

    # Start UI queue polling
    root.after(100, update_ui_from_queue)

    globals()["second_frame_global"] = second_frame
    # update globals for search
    globals()["header_frame_global"] = header_frame
    globals()["search_icon_global"] = search_icon_btn

    # Bind Ctrl+F after search bar + header exist
    root.bind("<Control-f>", lambda e: open_search_inline(tree, header_frame_global, search_icon_global))
    root.bind("<Control-F>", lambda e: open_search_inline(tree, header_frame_global, search_icon_global))

    root.mainloop()



def main(root_arg=None):
    home_page()



if __name__ == "__main__":
    home_page()

