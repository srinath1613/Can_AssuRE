import re, os, time, threading, subprocess, importlib, atexit, signal
import openpyxl
from queue import Queue
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import cantools, can
import config, gui
import sys
import copy
from functools import lru_cache



#------------------------
#Global objects
#-------------------
db = None # will load DBC once
bus = None
send_lock = threading.Lock()
stop_heartbeat = threading.Event()
stop_synchronized_event = threading.Event()   # 👈 new event for user signal thread
_synchronized_worker_started = False
signal_thread = None
user_signal_queue = Queue()
rx_queue = Queue()                     # for received signals
ui_queue = Queue()
excel_lock = threading.Lock()
shared_tx_signals = {}
shared_tx_lock = threading.Lock()
time_column_initialized = False
adb_device_connected = False
original_heartbeat_backup = None
user_send_signals_runtime = []
auto_send_running = False

# Fast cache for GUI Show Signals
frame_signal_cache = {}
property_value_cache = None
frame_list_cache = None



frameid_map = {}           # { frame_name: can_id_string }
_frameid_map_loaded = False

# shared memory for async ADB RX
shared_rx_latest = {}
shared_rx_lock = threading.Lock()
adb_worker_stop = threading.Event()

# -----------------------------
# Initialize Bus
# -----------------------------

def get_bus():
    global bus
    if bus is not None:
        return bus  # already initialized
    try:
        if config.Testing_device == "Linux":
            # Linux uses socketcan (PCAN-USB also appears as canX)
            bus = can.interface.Bus(
                channel=config.Peak_interface_name,
                bustype="socketcan"
            )
        elif config.Testing_device == "Windows":
            # Windows uses pcan (PEAK driver)
            bus = can.interface.Bus(
                channel=config.Peak_interface_name,
                bustype="pcan"
            )
        else:
            print(f"❌ Invalid Testing_device: {config.Testing_device}")
            bus = None

    except Exception as e:
        print(f"❌ CAN initialization failed for {config.Testing_device}: {e}")
        bus = None

    return bus

def shutdown_bus():
    """Shutdown and clear the global bus."""
    global bus
    try:
        stop_heartbeat.set()
        time.sleep(0.2)
        if bus is not None:
            try:
                bus.shutdown()
            except Exception as e:
                print("Error shutting down bus:", e)
            bus = None
    except Exception as e:
        print(f"⚠️ Error during CAN shutdown: {e}")

atexit.register(shutdown_bus)

#---------------------
#config file update
#---------------------
def copy_original_heartbeat_signal():
    global original_heartbeat_backup
    original_heartbeat_backup = tuple(copy.deepcopy(config.heart_beat_signals))

def update_vehicle_property_type(dbc_file_path ):
    db = cantools.database.load_file(dbc_file_path)
    signal_list = [sig.name for msg in db.messages for sig in msg.signals]
    propid_vehicletype_dict = {}

    # Pattern: Message__Signal_RX_V = property_id /* ... VehiclePropertyType:TYPE ... */
    pattern = re.compile(r"(\w+)__([\w_]+)_RX_V\s*=\s*(\d+).*VehiclePropertyType:([\w\d]+)")

    with open(config.type_h_file, "r", encoding="utf-8") as f:
        for line in f:
            match = pattern.search(line)
            if match:
                message_name = match.group(1)
                signal_name = match.group(2)
                prop_id_dec = int(match.group(3))
                prop_type = match.group(4)

                # Check if this signal exists in DBC signal list
                if signal_name in signal_list:
                    prop_id_hex = format(prop_id_dec, "x")  # decimal → hex without 0x
                    propid_vehicletype_dict[prop_id_hex] = prop_type
    return propid_vehicletype_dict


def check_auto_send_time_in_excel_update_config_file():
    """
    This will open vector.xls sheet and load sheet 0
    will check for the time provided in the sheet using "Time(ms) variable"
    time list taken from xl sheet and to update time list in config.py sheet
    """
    try:
        with excel_lock:
            wb = openpyxl.load_workbook(config.file_path)    #Load vector sheet
        wb._active_sheet_index = 0
        sheet = wb.active
        word = "Time (ms)"
        for row in sheet.iter_rows():      # finding time column and its entry till cell is blank
            for cell in row:
                if word in str(cell.value):
                    time_list = []
                    column = cell.column
                    row = cell.row + 1
                    while 1:
                        column_letter = openpyxl.utils.get_column_letter(column)
                        name = column_letter + str(row)
                        time_ms = sheet[name].value
                        if time_ms is None:
                            break
                        time_list.append(time_ms)
                        row += 1
        gui.update_config_file_runtime("timedelay", time_list ) #update config sheet list of time delay given on runtime
    finally:
        wb.close()


#----------------------
#make peak interface
#-----------------------

def check_peak_device_interface_name():
    """
    Detect CAN interface name and update config.Peak_interface_name.
    Works safely on both Linux and Windows.
    """
    try:
        if config.Testing_device == "Linux":
            # List only interfaces that start with can
            output = subprocess.run(
                "ls /sys/class/net | grep '^can'",
                shell=True,
                capture_output=True,
                text=True
            )
            interfaces = [i.strip() for i in output.stdout.splitlines() if i.strip()]
            if interfaces:
                # Choose first valid CAN interface (usually can0)
                peak_interface = interfaces[0]
                gui.update_config_file_runtime("Peak_interface_name", peak_interface)
                #print(f"Linux CAN interface detected: {peak_interface}")
            else:
                print("❌ No CAN interface found on Linux")
        elif config.Testing_device == "Windows":
            interface_list = can.detect_available_configs(interfaces="pcan")
            if not interface_list:
                print("❌ No PCAN interface detected on Windows")
                return
            # Pick the first config
            interface_dict = interface_list[0]
            # Safely extract channel
            channel = interface_dict.get("channel") or interface_dict.get("device")
            if channel:
                gui.update_config_file_runtime("Peak_interface_name", channel)
                #print(f"Windows PCAN channel detected: {channel}")
            else:
                print("❌ PCAN config missing 'channel' entry")
    except Exception as e:
        print("❌ Error detecting CAN interface:", e)


def make_can_interface_up():
    """
    By default, in linux can interface will be down
    using linux cmd we are making interface up
    so that peak interface can start send signal
    """
    if config.Testing_device == "Linux":
        password = config.Password
        cmd_list = [f'echo {password} | sudo -S ip link set {config.Peak_interface_name} down',
                    f'echo {password} | sudo -S ip link set {config.Peak_interface_name} type can',
                    f'echo {password} | sudo -S ip link set {config.Peak_interface_name} type can bitrate {config.Bit_rate}',
                    f'echo {password} | sudo -S ip link set {config.Peak_interface_name} up']     # writing the above cmd in Linux machine to make the interface up

        for cmd in cmd_list:
            try:
                os.system(cmd)
            except Exception as e:
                print(e)
    else:
        pass

def check_whether_can_interface_is_up():
    """
    Checks if the CAN interface (PEAK) is UP.
    Works safely on both Linux and Windows.
    No ADB commands inside this check!
    """

    # -------------------------
    # Linux: use socketcan
    # -------------------------
    if config.Testing_device == "Linux":
        try:
            result = subprocess.run(
                ["ip", "link", "show", config.Peak_interface_name],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL
            )
            return 1 if result.returncode == 0 else 0
        except Exception as e:
            print("Linux CAN check error:", e)
            return 0

    # -------------------------
    # Windows: use PCAN detection
    # -------------------------
    elif config.Testing_device == "Windows":
        try:
            configs = can.detect_available_configs(interfaces=['pcan'])
            return 1 if configs else 0
        except Exception as e:
            print("PCAN detection error:", e)
            return 0

    return 0

#------------------------
#Frame and signal utilities
#------------------------


def get_list_of_frames():
    importlib.reload(config)
    db_local = cantools.db.load_file(config.dbc_file_path)
    return [msg.name for msg in db_local.messages]

def get_list_of_signals(frame_name):
    db_local = cantools.db.load_file(config.dbc_file_path)
    for msg in db_local.messages:
        if msg.name == frame_name:
            return [sig.name for sig in msg.signals]
    return []

def load_vector_sheet():
    frame_name = get_list_of_frames()
    signal_list = []
    for frame in frame_name:
        signal_list += get_list_of_signals(frame)
    wb = openpyxl.load_workbook(config.file_path)
    wb._active_sheet_index = 0
    sheet = wb.active
    return signal_list, sheet


def _parse_type_h_property_map():
    """
    Parse config.type_h_file and return map: { signal_name: '0x....' }
    Pattern expected in types.h:
      Message__Signal_RX_V = 123456  /* ... VehiclePropertyType:TYPE ... */
    """
    prop_map = {}
    try:
        pattern = re.compile(r"([A-Za-z0-9_]+)__([A-Za-z0-9_]+)_RX_V\s*=\s*(\d+)", re.IGNORECASE)
        with open(config.type_h_file, "r", encoding="utf-8") as f:
            for line in f:
                m = pattern.search(line)
                if not m:
                    continue
                # message_name = m.group(1)   # unused here, we only need signal
                signal_name = m.group(2).strip()
                prop_id_dec = int(m.group(3))
                prop_id_hex = f"0x{prop_id_dec:x}"
                prop_map[signal_name] = prop_id_hex.lower()
    except Exception as e:
        # parsing failed or file missing -> return empty map
        print(f"⚠️ _parse_type_h_property_map error: {e}")
    return prop_map


def get_property_values():
    """
    Return list of property IDs aligned with the order of signals returned from
    load_vector_sheet() (signal_list order). Prefer types.h values, fallback to
    Excel search if types.h didn't contain the mapping.
    Returns: [ '0x21608350', '', '0x22f0...', ... ]  (lowercase, empty string if not found)
    """
    signal_list, sheet = load_vector_sheet()
    property_list = []

    # --- Try types.h first (fast, authoritative) ---
    try:
        typeh_map = _parse_type_h_property_map()
        if typeh_map:
            for sig in signal_list:
                # match exact name (case sensitive in types.h usually) and try fallback to insensitive
                prop = typeh_map.get(sig)
                if not prop:
                    # try case-insensitive match (some files vary)
                    for k, v in typeh_map.items():
                        if k.lower() == str(sig).lower():
                            prop = v
                            break
                property_list.append(prop or "")
            # If we found *some* entries (non-empty) return this mapping.
            # If all are empty, fall through to excel fallback.
            if any(property_list):
                # normalize to lowercase and return
                return [str(p).lower() if p else "N/A" for p in property_list]
    except Exception as e:
        print(f"⚠️ types.h lookup failed: {e}")

    # --- Fallback: scan Excel (existing logic, slightly hardened) ---
    try:
        property_list = []
        for word in signal_list:
            name = None
            # iterate rows and cells (sheet from load_vector_sheet)
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(word) in str(cell.value):
                        column = cell.column
                        row_idx = cell.row - 1
                        column_letter = get_column_letter(column)
                        name = f"{column_letter}{row_idx}"
                        break
                if name:
                    break
            if name:
                try:
                    prop = sheet[name].value
                    property_list.append(str(prop).lower() if prop is not None else "")
                except Exception:
                    property_list.append("")
            else:
                # not found in excel
                print(f"Warning: No match found for {word}")
                property_list.append("")
        return property_list
    except Exception as e:
        print(f"❌ Excel fallback get_property_values error: {e}")
        # final fallback: return empty placeholders
        return ["" for _ in signal_list]

def refresh_config_and_dbc():
    global db
    importlib.reload(config)
    print(f"🔁 Config reloaded: dbc_file_path={config.dbc_file_path}")
    if config.dbc_file_path and os.path.exists(config.dbc_file_path):
        db = cantools.db.load_file(config.dbc_file_path)
        print(f"✅ DBC loaded successfully: {config.dbc_file_path}")
    else:
        print("⚠️ No valid DBC path found.")

    # preload vector sheet cache (frame -> can id)
    frameid_map.clear()
    _frameid_map_loaded = False
    preload_vector_sheet()

def load_vector_sheet_frameid(frame_name):
    """
    Fetch CAN ID for a frame from preloaded frameid_map.
    If the cache is not loaded yet, try to preload it once (thread-safe).
    """
    global _frameid_map_loaded

    # Ensure cache is loaded
    if not _frameid_map_loaded:
        preload_vector_sheet()

    # Direct lookup (no file I/O)
    can_id = frameid_map.get(str(frame_name).strip())
    if can_id:
        return str(can_id).strip()
    else:
        # fallback: attempt one locked read of workbook (rare)
        with excel_lock:
            try:
                if not os.path.exists(config.file_path):
                    return None
                wb = openpyxl.load_workbook(config.file_path, data_only=True)
                sheet = wb.worksheets[1]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 2:
                        f = row[0]
                        cid = row[1]
                        if str(f).strip() == str(frame_name).strip():
                            wb.close()
                            # update cache so future lookups are fast
                            frameid_map[str(f).strip()] = str(cid).strip()
                            return str(cid).strip()
                wb.close()
            except Exception as e:
                print(f"❌ Error reading vector sheet fallback: {e}")
        return None


#------------------------------------
#Prepare Signal Dictionary from excel
#-----------------------------------

def send_signal_values():
    #
   # Returns dict: {signal_name: {timedelay: value}}
    #
    signal_list, sheet = load_vector_sheet()
    signal_value_dict = {}
    for word in signal_list:
        for row in sheet.iter_rows():
            for cell in row:
                if word in str(cell.value):
                    value = {}
                    for i in range(1, len(config.timedelay)+1):
                        col = cell.column
                        row_idx = cell.row + i
                        column_letter = get_column_letter(col)
                        name = f"{column_letter}{row_idx}"
                        sig_value = sheet[name].value or 0
                        value[config.timedelay[i-1]] = sig_value
                    signal_value_dict[word] = value
    return signal_value_dict

# Track if Results sheet was initialized (Time + TX copied)
time_column_initialized = False

# ================================================================
# GLOBAL MEMORY CACHE (Excel loaded only once)
# ================================================================
EXCEL_FAST_CACHE = {
    "wb": None,
    "sheet": None,
    "time_row": {},      # {1000: 4, 2000: 5, ...}
    "tx_col": {},        # {signal_name: column}
    "rx_col": {},        # {signal_name: column}
    "initialized": False,
    "tx_initialized": False
}


# ================================================================
# 1️⃣ INITIALIZE SHEET STRUCTURE (Time + TX Copy) — RUN ONCE
# ================================================================
def initialize_results_sheet_structure():
    """
    Creates:
    - Time(ms) column
    - TX copy from CAN_Signals to Results
    This executes only ONCE.
    """

    if EXCEL_FAST_CACHE["tx_initialized"]:
        return

    file_path = config.file_path
    if not os.path.exists(file_path):
        print("❌ Excel file missing")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Results"]
    can_sheet = wb["CAN_Signals"]

    # ------------ Add Time(ms) header -----------
    sheet.cell(row=2, column=1).value = "Time (ms)"

    # ------------ Write time rows -----------
    if hasattr(config, "timedelay"):
        for i, t in enumerate(config.timedelay):
            sheet.cell(row=4 + i, column=1).value = int(t)
        print("⏱ Time(ms) column written")

    # ------------ Copy TX -----------
    print("📋 Copying TX values from CAN_Signals → Results ...")

    # build mapping CAN_Signals → column
    can_map = {
        str(can_sheet.cell(row=2, column=c).value).strip(): c
        for c in range(2, can_sheet.max_column + 1)
        if can_sheet.cell(row=2, column=c).value
    }

    # build mapping Results → column
    result_map = {}
    for c in range(2, sheet.max_column + 1):
        sig = sheet.cell(row=2, column=c).value
        if sig:
            result_map[str(sig).strip()] = c

    # copy data
    for sig, can_col in can_map.items():
        if sig not in result_map:
            continue
        res_col = result_map[sig]

        # copy downward (row 3 in CAN → row 4 in Results)
        for r in range(3, can_sheet.max_row + 1):
            sheet.cell(row=r + 1, column=res_col).value = can_sheet.cell(row=r, column=can_col).value

    print("💾 TX columns copied")

    wb.save(file_path)
    print("✔ Initial Results sheet structure saved")

    EXCEL_FAST_CACHE["tx_initialized"] = True



# ================================================================
# 2️⃣ LOAD EXCEL INTO MEMORY + BUILD LOOKUP TABLES (RUN ONCE)
# ================================================================
def init_fast_excel():
    """
    Loads workbook once and prepares:
    - time_row mapping
    - rx_col mapping
    """

    if EXCEL_FAST_CACHE["initialized"]:
        return

    file_path = config.file_path

    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Results"]

    EXCEL_FAST_CACHE["wb"] = wb
    EXCEL_FAST_CACHE["sheet"] = sheet

    # -------- time_row mapping --------
    time_map = {}
    for r in range(4, sheet.max_row + 1):
        val = sheet.cell(row=r, column=1).value
        if val is not None:
            time_map[int(val)] = r

    EXCEL_FAST_CACHE["time_row"] = time_map

    # -------- rx_col mapping --------
    rx_map = {}
    for c in range(2, sheet.max_column + 1):
        subheader = sheet.cell(row=3, column=c).value
        if subheader and "rx" in str(subheader).lower():
            sig = sheet.cell(row=2, column=c - 1).value
            if sig:
                rx_map[str(sig).strip()] = c

    EXCEL_FAST_CACHE["rx_col"] = rx_map

    print("⚡ FAST Excel initialized (memory mode)")
    EXCEL_FAST_CACHE["initialized"] = True



# ================================================================
# 3️⃣ FAST RX UPDATE (NO SAVE, NO RELOAD)
# ================================================================
def fast_update_excel_rx(time_ms, adb_rx_values):
    """
    Super-fast RX update:
    - No workbook reload
    - No save
    - O(1) column lookup
    """

    init_fast_excel()

    sheet = EXCEL_FAST_CACHE["sheet"]
    time_row = EXCEL_FAST_CACHE["time_row"].get(time_ms)

    if not time_row:
        print(f"⚠ Time {time_ms} not found")
        return

    rx_col_map = EXCEL_FAST_CACHE["rx_col"]
    updated = []

    for sig, val in adb_rx_values.items():
        if sig in rx_col_map:
            col = rx_col_map[sig]
            sheet.cell(row=time_row, column=col).value = val
            updated.append(sig)

    print(f"📥 RX updated in memory for {len(updated)} signals @ {time_ms} ms")



# ================================================================
# 4️⃣ SAVE ONCE AT END
# ================================================================
def save_fast_excel():
    if EXCEL_FAST_CACHE["wb"]:
        EXCEL_FAST_CACHE["wb"].save(config.file_path)
        print("💾 FAST Excel saved (one-time)")


def Update_Value_Tx_column_in_UI(signal_dict_msg):
    try:
        ui_queue.put(signal_dict_msg)
        #gui.root.after_idle(gui.update_ui_from_queue)  # immediately process next queue item
        print(f"📤 UI TX Update Queued: {signal_dict_msg}")
        time.sleep(0.01)  # small yield so GUI can breathe
    except Exception as e:
        print(f"⚠️ Failed to queue UI TX update: {e}")


def read_property_ids_from_excel(file_path):
    """
    Read CAN_Signals sheet and return:
    property_id_map = { signal_name: property_id }
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb["CAN_Signals"]

        property_id_map = {}

        # Row 1 -> Property ID
        # Row 2 -> Signal Name
        # Columns start from 2 because column A is "Property ID"
        col = 2
        while ws.cell(row=2, column=col).value:
            signal_name = ws.cell(row=2, column=col).value
            property_id = ws.cell(row=1, column=col).value
            if signal_name and property_id:
                property_id_map[signal_name] = property_id
            col += 1
        return property_id_map

    except Exception as e:
        print(f"❌ Error while reading Excel: {e}")

def read_signal_values_from_excel(file_path):
    """
    Reads CAN_Signals sheet and returns dictionary:
    {
        signal_name: {
            time1: val1,
            time2: val2,
            ...
        }
    }
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb["CAN_Signals"]

        signal_values = {}

        # --- Read all signal names (Row 2, columns 2..n)
        col = 2
        signal_names = []
        while ws.cell(row=2, column=col).value:
            signal_names.append(ws.cell(row=2, column=col).value)
            col += 1

        # --- Read all time values (Column A starting from row 3)
        row = 3
        times = []
        while ws.cell(row=row, column=1).value is not None:
            times.append(ws.cell(row=row, column=1).value)
            row += 1

        # --- Build the dictionary
        for idx, signal_name in enumerate(signal_names, start=2):
            time_value_map = {}
            for r in range(3, 3 + len(times)):
                time_val = ws.cell(row=r, column=1).value
                cell_val = ws.cell(row=r, column=idx).value

                # If empty → treat as 0
                if cell_val is None:
                    cell_val = 0

                time_value_map[time_val] = cell_val

            signal_values[signal_name] = time_value_map

        return signal_values

    except Exception as e:
        print(f"❌ Error reading signal values: {e}")
        return {}

@lru_cache(maxsize=1)
def cached_signal_dict():
    return read_signal_values_from_excel(config.file_path)

@lru_cache(maxsize=1)
def cached_property_map():
    return read_property_ids_from_excel(config.file_path)


def synchronized_signal_worker():
    global db

    try:
        # ====================================================
        # 1) Load DBC only once
        # ====================================================
        if db is None:
            if not config.dbc_file_path or not os.path.exists(config.dbc_file_path):
                print("❌ Invalid DBC path:", config.dbc_file_path)
                return

            db = cantools.database.load_file(config.dbc_file_path)
            print(f"✅ Loaded DBC: {config.dbc_file_path}")

        # Build fast lookup → {frame_name: frame_id}
        dbc_frameid_map = {msg.name: msg.frame_id for msg in db.messages}


        # ====================================================
        # 2) Load Excel only once
        # ====================================================
        excel_path = config.file_path
        initialize_results_sheet_structure()
        property_map = cached_property_map()
        signal_dict = cached_signal_dict()

        # Validate timedelay list
        if not hasattr(config, "timedelay") or not config.timedelay:
            print("⚠️ No timedelay values in config")
            return

        print("⏱ Timedelay sequence:", config.timedelay)



        # ====================================================
        # MAIN LOOP — One cycle per delay
        # ====================================================
        heartbeat_frames = {hb["frame_name"] for hb in config.heart_beat_signals}
        first_loop = True
        for td in config.timedelay:
            if stop_synchronized_event.is_set():
                print("🛑 Sync worker stop requested — exiting loop")
                break
            print(f"\n🕒 Starting cycle for {td} ms")
            time.sleep(td / 1000.0)

            # ====================================================
            # 3) Build frame_groups with CAN-ID + signals
            # ====================================================
            # Format:
            # frame_groups = {
            #     "Batt_Sts_Info": {
            #         "can_id": 0x12d,
            #         "signals": { "Display_SoC": 99, "Batt_Curr": 10 }
            #     }
            # }

            frame_groups = {}

            for sig_name, td_values in signal_dict.items():

                frame = get_frame_from_signal(sig_name)
                if not frame:
                    continue

                # value for this delay
                val = td_values.get(td, td_values[max(td_values.keys())])

                if frame not in frame_groups:
                    frame_groups[frame] = {
                        "can_id": dbc_frameid_map.get(frame),
                        "signals": {}
                    }

                frame_groups[frame]["signals"][sig_name] = val



            # ====================================================
            # 4) Build user_send_signals + update heartbeat signals
            # ====================================================
            new_user_send_signals = []

            for frame_name, frame_info in frame_groups.items():

                frame_id = frame_info["can_id"]
                sig_map = frame_info["signals"]

                if not frame_id:
                    print(f"⚠️ Missing CAN-ID for frame {frame_name}, skipping…")
                    continue


                # ------------------------------
                # HEARTBEAT FRAME → update only heartbeat list
                # ------------------------------
                if frame_name in heartbeat_frames:

                    for hb in config.heart_beat_signals:
                        if hb["frame_name"] == frame_name:
                            # update only matching signal names
                            for sig, val in sig_map.items():
                                if sig in hb["signals"]:
                                    hb["signals"][sig] = val
                    continue


                # ------------------------------
                # NORMAL FRAMES → user_send_signals entry
                # ------------------------------
                new_user_send_signals.append({
                    "frame_name": frame_name,
                    "can_id": frame_id,
                    "signals": sig_map
                })



            # ====================================================
            # 5) Update config runtime once per delay
            # ====================================================
            gui.update_config_file_runtime("heart_beat_signals", config.heart_beat_signals)
            #gui.update_config_file_runtime("user_send_signals", new_user_send_signals)
            importlib.reload(config)
            user_send_signals_runtime = new_user_send_signals

            print(f"🟩 Updated signals for {td} ms")



            # ====================================================
            # 6) Run TX + RX Threads
            # ====================================================

            def update_tx_ui():
                tx_dict = {
                    "tx_update": {
                        sig: val
                        for f in frame_groups.values()
                        for sig, val in f["signals"].items()
                    }
                }
                Update_Value_Tx_column_in_UI(tx_dict)
                print(f"📤 UI Tx updated for {td} ms")

            def validate_and_update_rx():
                nonlocal first_loop
                global shared_rx_latest

                # --- ON FIRST LOOP: restart adb worker to remove stale internal state ---
                if first_loop:
                    try:
                        stop_adb_worker()
                    except Exception:
                        pass
                    time.sleep(0.05)
                    try:
                        start_adb_worker()
                    except Exception as e:
                        print("⚠️ Error restarting ADB worker:", e)

                # --- Always clear old snapshot before waiting for fresh data ---
                try:
                    with shared_rx_lock:
                        # clear previous snapshot so we can detect a new one
                        if isinstance(shared_rx_latest, dict):
                            shared_rx_latest.clear()
                        else:
                            shared_rx_latest = {}
                except Exception:
                    try:
                        shared_rx_latest = {}
                    except Exception:
                        pass

                # --- Wait for a fresh snapshot (poll) ---
                wait_timeout = 1.5  # seconds — tune if your device is slower
                poll_interval = 0.05
                end_time = time.time() + wait_timeout
                got_snapshot = False

                while time.time() < end_time:
                    with shared_rx_lock:
                        if shared_rx_latest:  # non-empty dict means fresh data arrived
                            got_snapshot = True
                            latest = dict(shared_rx_latest)
                            break
                    time.sleep(poll_interval)

                if not got_snapshot:
                    # fallback: attempt to use whatever is present (may be empty)
                    with shared_rx_lock:
                        latest = dict(shared_rx_latest) if shared_rx_latest else {}
                    #print(f"⚠️ No fresh adb snapshot within {wait_timeout}s — using fallback (may be empty)")

                # --- Update Excel + UI using the fresh/latest snapshot ---
                try:
                    # lock excel if you do concurrent writes elsewhere
                    with excel_lock:
                        fast_update_excel_rx(int(td), latest)
                    ui_queue.put(latest)
                    print(f"📡 Async RX updated for {td} ms (got_snapshot={got_snapshot})")
                except Exception as e:
                    print(f"❌ Async RX update failed for {td} ms:", e)

            def delayed_rx_start(delay_ms=500):
                time.sleep(delay_ms / 1000.0)
                validate_and_update_rx()

            tx_thread = threading.Thread(target=update_tx_ui)
            rx_thread = threading.Thread(target=delayed_rx_start, args=(500,))

            tx_thread.start()
            rx_thread.start()
            tx_thread.join()
            rx_thread.join()

            print(f"✔ Completed Tx/Rx cycle {td} ms")
            first_loop = False

        print("\n🏁 ALL cycles done successfully!")
        save_fast_excel()

    except Exception as e:
        print(f"❌ Worker crashed:", e)

    finally:
        print("🛑 Worker exiting…")
        stop_synchronized_worker()




def start_synchronized_worker():
    global _synchronized_worker_started, stop_synchronized_event, auto_send_running
    if not _synchronized_worker_started:
        stop_synchronized_event.clear()
        threading.Thread(target=synchronized_signal_worker, daemon=True).start()
        _synchronized_worker_started = True
        auto_send_running = True
        print("✅ backend: synchronized worker started")
    else:
        print("⚙️ Worker already running.")

def stop_synchronized_worker():
    """
    Stop only the synchronized signal worker thread safely.
    """
    global _synchronized_worker_started, stop_synchronized_event, auto_send_running
    if _synchronized_worker_started:
        print("🛑 Stopping synchronized worker...")
        stop_synchronized_event.set()   # signal the worker to stop
        _synchronized_worker_started = False
        auto_send_running = False
        gui.send.config(state=gui.tk.NORMAL)
    else:
        print("ℹ️ No synchronized worker active.")




# -----------------------------
# ADB Poll Worker
# -----------------------------
def build_canid_to_signalname_from_excel():
    """
    Reads CAN ID → signal names mapping from the vector Excel sheet (first page).
    Expected format:
        Row 1: Property ID | 0x21608350 | 0x21608351 | ...
        Row 2: Time (ms)   | MCU_DC_Curr | MCU_Temp  | ...
    """
    try:
        with excel_lock:
            wb = openpyxl.load_workbook(config.file_path, data_only=True)
        #wb = openpyxl.load_workbook(config.file_path)
        sheet = wb.worksheets[0]  # First sheet

        canid_to_signalname = {}

        # Find header row (Property ID)
        header_row = [cell.value for cell in sheet[1] if cell.value]
        if not header_row:
            print("⚠️ No header row found in Excel sheet!")
            return {}

        # The next row should contain signal names
        signal_row = [cell.value for cell in sheet[2] if cell.value]

        for col in range(1, len(header_row)):
            can_id_cell = header_row[col]
            signal_name = signal_row[col] if col < len(signal_row) else None

            if can_id_cell and signal_name:
                can_id = str(can_id_cell).strip().replace("0x", "").replace("’", "").replace("‘", "").lower()
                canid_to_signalname.setdefault(can_id, []).append(signal_name)

        #print(f"✅ Built CAN ID → Signal map: {canid_to_signalname}")
        return canid_to_signalname

    except Exception as e:
        print(f"⚠️ Error reading Excel for CAN ID mapping: {e}")
        return {}
    finally:
        wb.close()
# ---------------------------------------------------------
# Cache CAN-ID → list of signals (runs only once)
# ---------------------------------------------------------
@lru_cache(maxsize=1)
def cached_canid_to_signalname():
    return build_canid_to_signalname_from_excel()


# ---------------------------------------------------------
# Cache property list (runs only once)
# ---------------------------------------------------------
@lru_cache(maxsize=1)
def cached_property_list():
    return get_property_values()


# ---------------------------------------------------------
# COMPILED REGEX cache (one per field type)
# ---------------------------------------------------------
REGEX_CACHE = {}


def fast_get_regex(prop_key, field_name):
    """
    Return compiled regex for faster repeated matching.
    Stored in REGEX_CACHE to avoid re-compilation.
    """
    key = (prop_key, field_name)
    if key not in REGEX_CACHE:
        pattern = rf"Property:\s*0x{prop_key}[\s\S]*?{field_name}:\s*\[([^\]]*)\]"
        REGEX_CACHE[key] = re.compile(pattern, re.IGNORECASE)
    return REGEX_CACHE[key]


# ---------------------------------------------------------
# FAST, OPTIMIZED validate_vhal_layer
# ---------------------------------------------------------
def validate_vhal_layer():
    property_list = cached_property_list()
    canid_to_signalname = cached_canid_to_signalname()

    carservice = {}

    #print("📡 Fetching all VHAL properties in ONE adb call…")

    # -----------------------------------------------------
    # 1️⃣  Run single ADB dump
    # -----------------------------------------------------
    try:
        cmd = "adb shell dumpsys car_service get-property-value"
        result = subprocess.run(
            cmd, shell=True, capture_output=True, text=True, timeout=10
        )
        full_output = result.stdout

    except subprocess.TimeoutExpired:
        print("❌ ADB timeout")
        return fill_device_not_found(property_list, canid_to_signalname)

    except Exception as e:
        print(f"❌ ADB failure: {e}")
        return fill_device_not_found(property_list, canid_to_signalname)

    if not full_output.strip():
        print("⚠️ Empty ADB response")
        return fill_device_not_found(property_list, canid_to_signalname)

    # -----------------------------------------------------
    # 2️⃣  Parse all properties in memory FAST
    # -----------------------------------------------------
    for prop_id in property_list:

        # Pre-optimized string conversion
        raw = str(prop_id).lower().replace("0x", "")
        prop_key = raw.replace("’", "").replace("‘", "")  # safe

        prop_type = config.Vehicle_propID_type.get(prop_key, "FLOAT")
        field_name = {
            "INT32": "int32Values",
            "INT64": "int64Values",
            "BYTES": "bytes",
            "STRING": "string",
        }.get(prop_type, "floatValues")

        # Precompiled regex lookup
        regex = fast_get_regex(prop_key, field_name)
        match = regex.search(full_output)

        value = match.group(1).strip() if match else "Not found"

        # Map CAN-ID → signals
        signal_names = canid_to_signalname.get(raw, [])

        for sig in signal_names:
            carservice[sig] = value

    # -----------------------------------------------------
    # 3️⃣  Push values to UI queue
    # -----------------------------------------------------
    ui_queue.put(carservice)
    #print("📡 Parsed VHAL:", carservice)

    return carservice


# ---------------------------------------------------------
# Helper for device-not-found fast return
# ---------------------------------------------------------
def fill_device_not_found(property_list, canid_to_signalname):
    result = {}

    for prop_id in property_list:
        can_id = str(prop_id).lower().replace("0x", "")
        signals = canid_to_signalname.get(can_id, [])
        for sig in signals:
            result[sig] = "Device not found"

    ui_queue.put(result)
    return result

# -----------------------------
# Helper function
# -----------------------------
def get_frame_from_signal(signal_name):
    global db
    if db is None:
        if not config.dbc_file_path or not os.path.exists(config.dbc_file_path):
            print("❌ DBC file not loaded in config")
            return None
        db = cantools.database.load_file(config.dbc_file_path)
    for frame in get_list_of_frames():
        if signal_name in get_list_of_signals(frame):
            return frame
    return None


def preload_vector_sheet():
    """
    Load the 2nd sheet (frame -> CAN ID) into memory (frameid_map).
    This avoids repeatedly opening the .xlsx from multiple threads.
    """
    global frameid_map, _frameid_map_loaded

    # If already loaded, just return
    if _frameid_map_loaded:
        return

    # Acquire excel_lock to avoid concurrent workbook access
    with excel_lock:
        try:
            if not os.path.exists(config.file_path):
                #print(f"⚠️ preload_vector_sheet: file not found: {config.file_path}")
                frameid_map = {}
                _frameid_map_loaded = True
                return

            wb = openpyxl.load_workbook(config.file_path, data_only=True, read_only=True)
            # 2nd sheet expected to contain FrameName | CAN ID
            if len(wb.sheetnames) < 2:
                #print("⚠️ preload_vector_sheet: workbook has less than 2 sheets")
                wb.close()
                frameid_map = {}
                _frameid_map_loaded = True
                return

            sheet = wb.worksheets[1]  # second sheet
            temp_map = {}
            # assume header at row 1, data from row 2
            for r in range(2, sheet.max_row + 1):
                frame_cell = sheet.cell(row=r, column=1).value
                canid_cell = sheet.cell(row=r, column=2).value
                if frame_cell and canid_cell:
                    try:
                        frame_name = str(frame_cell).strip()
                        can_id = str(canid_cell).strip()
                        temp_map[frame_name] = can_id
                    except Exception:
                        continue

            wb.close()
            frameid_map = temp_map
            _frameid_map_loaded = True
            #print(f"✅ preload_vector_sheet: loaded {len(frameid_map)} frames")
        except Exception as e:
            print(f"❌ preload_vector_sheet error: {e}")
            frameid_map = {}
            _frameid_map_loaded = True


#------------------------
#Heart Beat signals
#--------------------------
def get_adb_devices():
    """Return a list of connected ADB device serial numbers"""
    try:
        output = subprocess.check_output(["adb", "devices"], stderr=subprocess.STDOUT).decode()
        lines = output.strip().split("\n")[1:]  # skip the first header line
        devices = [line.split()[0] for line in lines if "\tdevice" in line]
        return devices
    except subprocess.CalledProcessError as e:
        print("Error checking ADB devices:", e.output.decode())
        return []
    except FileNotFoundError:
        print("ADB not found — make sure it's installed and in PATH.")
        return []

def monitor_adb_device():
    global adb_device_connected
    while True:
        devices = get_adb_devices()
        adb_device_connected = len(devices) > 0
        time.sleep(1)    # check every 1 sec ONLY

threading.Thread(target=monitor_adb_device, daemon=True).start()

def check_device_mode():
    try:
        if not adb_device_connected:
            return {}
        # Read raw output from ADB
        cmd = "adb shell dumpsys car_service get-property-value"
        output = subprocess.check_output(cmd, shell=True, text=True)
        # Normalize property ID to hex without "0x"
        prop_id = config.Drive_mode_prop_ID.lower().replace("0x", "").strip()
        # Find the block that contains the property
        pattern = rf"Property:\s*0x{prop_id}[\s\S]*?int32Values:\s*\[([^\]]*)\]"
        match = re.search(pattern, output, re.IGNORECASE)
        # Extract mode value
        mode = int(match.group(1).strip()) if match else -1
        # Mode mapping
        modes = {
            0: "VL1_SNA",
            1: "Sleep",
            2: "Awake",
            3: "Standby",
            4: "Drive",
            5: "Charge"
        }
        print(f"Device in {modes.get(mode, 'Unknown')} mode")
        return 1 if mode == 4 else 0
    except subprocess.CalledProcessError:
        print("Error: Failed to run adb command")
        return 0
    except Exception as e:
        print(f"Error: {e}")
        return 0


def send_heartbeat_frame(db, local_bus,frame_id_int,frame_name,signal_dict):
    try:
        message = db.get_message_by_frame_id(frame_id_int)
        message_data = db.encode_message(frame_name, signal_dict)
        frame = can.Message(arbitration_id=frame_id_int, data=message_data, is_extended_id=message.is_extended_frame)
        with send_lock:
            local_bus.send(frame)
    except cantools.database.errors.EncodeError as e:
        print(f"Encode error for frame {frame_name}: {e}")
    except can.CanError as e:
        print(f"CAN send error for frame {frame_name}: {e}")


def Send_Heart_beat_signal_continously_in_backgorund():
        try:
            db = cantools.db.load_file(config.Heart_beat_dbc)
            local_bus = get_bus()
            heartbeat_period = 0.08
            next_time = time.perf_counter()
            if local_bus is None:
                print("❌ Heartbeat: CAN bus not available on start")
                return
            config_path = config.__file__
            last_modified = os.path.getmtime(config_path)

            while not stop_heartbeat.is_set():
                start = time.perf_counter()
                # ✅ Reload config only if the file has been updated
                current_mtime = os.path.getmtime(config_path)
                if current_mtime != last_modified:
                    importlib.reload(config)
                    last_modified = current_mtime
                    #print("♻️ Config file reloaded due to update")

                if check_whether_can_interface_is_up() and adb_device_connected:

                    # ✅ Merge heartbeat + user signals dynamically
                    all_signals = list(config.heart_beat_signals)
                    if "user_send_signals_runtime" in globals():
                        all_signals.extend(user_send_signals_runtime)

                    for sig in all_signals:
                        try:
                            frame_id = sig["can_id"]
                            frame_id_int = int(frame_id, 16) if isinstance(frame_id, str) else int(frame_id)
                            send_heartbeat_frame(
                                db, local_bus, frame_id_int, sig["frame_name"], sig["signals"]
                            )
                            time.sleep(0.001)
                        except Exception as e:
                            print(f"⚠️ Failed to send frame {sig.get('frame_name', '?')}: {e}")

                else:
                    if not check_whether_can_interface_is_up():
                        print("❌ PEAK interface is down")
                    if not adb_device_connected:
                        print("❌ No ADB device detected")
                    if not check_device_mode():
                        print("❌ Device not in drive mode")
                next_time += heartbeat_period
                sleep_time = next_time - time.perf_counter()
                if sleep_time > 0:
                    steps = int(sleep_time / 0.01)
                    for _ in range(max(1, steps)):
                        if stop_heartbeat.is_set():
                            return
                        time.sleep(0.01)
                else:
                    # if we're running late, resync
                    next_time = time.perf_counter()
                #check_device_mode()

        except Exception as error:
                print("Unable to send heartbeat signals"+ str(error))


def cleanup_on_exit():
    """Stop all background threads and close CAN interface safely."""
    print("🧹 Cleaning up... stopping heartbeat and closing CAN interface")
    try:
        gui.update_config_file_runtime("user_send_signals", [])
        gui.update_config_file_runtime("Vehicle_propID_type", {})
        gui.update_config_file_runtime("heart_beat_signals", list(original_heartbeat_backup))
        importlib.reload(config)
    except Exception as e:
        print(f"⚠️ Failed to clear user_send_signals: {e}")
    try:
        stop_heartbeat.set()   # stop background loop if it checks this
        if 'bus' in globals() and bus is not None:
            try:
                bus.shutdown()
                print("✅ CAN bus shutdown successful")
            except Exception as e:
                print(f"⚠️ Error shutting down bus: {e}")
    except Exception as e:
        print("⚠️ Cleanup failed:", e)

    # Optionally bring the interface down in Linux
    if config.Testing_device == "Linux":
        password = config.Password
        os.system(f'echo {password} | sudo -S ip link set {config.Peak_interface_name} down')
        print(f"📴 Interface {config.Peak_interface_name} set down")
    global shared_rx_latest

    # ======================================================
    # 🚀 1) STOP ADB BACKGROUND WORKER
    # ======================================================
    try:
        adb_worker_stop.set()  # ask thread to exit
        #print("🛑 ADB worker stop signal sent")
    except Exception as e:
        print("⚠️ Failed sending stop to ADB worker:", e)

    # reset last cached ADB values
    try:
        shared_rx_latest = {}
    except:
        pass

    # ======================================================
    # 🚀 2) STOP SYNCHRONIZED WORKER
    # ======================================================
    try:
        stop_synchronized_worker()
    except Exception as e:
        print("⚠️ Failed stopping synchronized worker:", e)

    # ======================================================
    # 🚀 3) CLEAR FAST-EXCEL CACHE
    # ======================================================
    try:
        global EXCEL_FAST_CACHE
        EXCEL_FAST_CACHE.clear()
        EXCEL_FAST_CACHE.update({
            "wb": None,
            "sheet": None,
            "time_row": {},
            "tx_col": {},
            "rx_col": {},
            "initialized": False,
            "tx_initialized": False
        })
        print("🧹 Excel fast-cache cleared")
    except Exception as e:
        print("⚠️ Failed clearing Excel cache:", e)

    # ======================================================

# Register it for any exit condition
atexit.register(cleanup_on_exit)

def signal_handler(sig, frame):
    print("\n🛑 Keyboard interrupt detected — stopping heartbeat...")
    cleanup_on_exit()
    sys.exit(0)   # safer than os._exit()

# Always works
signal.signal(signal.SIGINT, signal_handler)

# Only register SIGTERM if available (Linux/Unix)
if hasattr(signal, "SIGTERM"):
    signal.signal(signal.SIGTERM, signal_handler)


signal_thread = threading.Thread(target=Send_Heart_beat_signal_continously_in_backgorund, daemon=True)
signal_thread.start()


def monitor_peak_device():
    """
    Continuously checks if PEAK device is connected or removed every 1 sec.
    If interface is down, it automatically tries to bring it UP again.
    Sends status to GUI via ui_queue.
    """
    last_status = None

    while True:
        try:
            status = check_whether_can_interface_is_up()  # 1 or 0

            if status == 0:
                # Interface DOWN → Try to auto-recover
                try:
                    print("🔧 PEAK interface down — trying to bring it UP...")
                    if config.Testing_device == "Linux":
                        make_can_interface_up()
                    time.sleep(0.5)  # Give system time

                    # Re-check after repair attempt
                    status = check_whether_can_interface_is_up()
                except Exception as e:
                    print("⚠️ Error while trying to bring interface up:", e)

            # Status changed → Send to GUI
            if status != last_status:
                if status == 1:
                    ui_queue.put({"peak_status": "connected"})
                    print("🔌 PEAK Device Connected")
                else:
                    ui_queue.put({"peak_status": "disconnected"})
                    print("❌ PEAK Device Removed / Not Ready")

                last_status = status

        except Exception as e:
            print(f"⚠️ PEAK monitor error: {e}")

        time.sleep(1)  # check every second


threading.Thread(target=monitor_peak_device, daemon=True).start()




def adb_background_worker():
    global shared_rx_latest, _adb_worker_started

    while not adb_worker_stop.is_set():
        try:
            adb_rx = validate_vhal_layer()
            with shared_rx_lock:
                shared_rx_latest = dict(adb_rx)
        except Exception as e:
            print("⚠️ ADB worker error:", e)

        #for _ in range(30):
        #    if adb_worker_stop.is_set():
        #        break
        #    time.sleep(0.01)
        signal_count = len(cached_signal_dict())

        if signal_count > 300:
            time.sleep(0.3)  # slower
        elif signal_count > 150:
            time.sleep(0.15)
        else:
            time.sleep(0.01)

    # Worker finished — mark state
    with adb_worker_lock:
        _adb_worker_started = False

    print("🛑 ADB worker exited cleanly")


# ======================================================
# ADB WORKER LIFECYCLE MANAGER
# ======================================================
adb_worker_thread = None
adb_worker_lock = threading.Lock()
_adb_worker_started = False


def start_adb_worker():
    """
    Start ADB background worker exactly once.
    Safe to call multiple times.
    """
    global adb_worker_thread, _adb_worker_started
    with adb_worker_lock:
        if _adb_worker_started:
            return False   # already running

        adb_worker_stop.clear()
        adb_worker_thread = threading.Thread(
            target=adb_background_worker,
            daemon=True
        )
        adb_worker_thread.start()
        _adb_worker_started = True
        print("✅ backend: ADB worker started")
        return True


def stop_adb_worker():
    """
    Stop ADB worker safely.
    """
    global adb_worker_thread, _adb_worker_started
    with adb_worker_lock:
        if not _adb_worker_started:
            return False

        adb_worker_stop.set()
        adb_worker_thread = None
        _adb_worker_started = False
        print("🛑 backend: ADB worker stop requested")
        return True


def is_adb_worker_running():
    return _adb_worker_started









